#!/usr/bin/env python3
"""
excel_scrape.py

Download activity-log Excel files for the primary 36 centers.

Source CSV: center_id_data.csv (expects columns: center_id, symbol, parent_id, status_cd)
Download URL pattern:
  https://iac.university/center/{symbol}/internal/activity-log/download

Environment variables (optional):
  IAC_EMAIL / IAC_PASSWORD  (required, loaded from .env if present)
  OUT_DIR=downloads/activity_logs
  CENTER_CSV=center_id_data.csv
	SQL_PATH=004_assessment_data.sql
  MAX_CENTERS=36
  REQUEST_DELAY_SEC=0.25

Usage:
  python excel_scrape.py
"""

import csv
import os
import re
import sys
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, List, Optional, Sequence

import requests
from bs4 import BeautifulSoup
from dotenv import load_dotenv
from openpyxl import load_workbook

LOGIN_URL = "https://iac.university/login"
INTERNAL_PROBE_URL = "https://iac.university/center/AS/internal"
DOWNLOAD_URL = "https://iac.university/center/{symbol}/internal/activity-log/download"

load_dotenv()

OUT_DIR = os.getenv("OUT_DIR", "downloads/activity_logs").strip()
CENTER_CSV = os.getenv("CENTER_CSV", "center_id_data.csv").strip()
SQL_PATH = os.getenv("SQL_PATH", "004_assessment_data.sql").strip()
MAX_CENTERS = int(os.getenv("MAX_CENTERS", "36"))
REQUEST_DELAY_SEC = float(os.getenv("REQUEST_DELAY_SEC", "0.25"))


class DownloadError(Exception):
	pass


@dataclass
class Center:
	center_id: int
	symbol: str
	name: str


def build_session() -> requests.Session:
	s = requests.Session()
	s.headers.update({
		"User-Agent": "Mozilla/5.0 (compatible; IAC-Excel-Downloader/1.0)",
		"Accept-Language": "en-US,en;q=0.9",
	})
	return s


def _extract_csrf_from_html(html: str) -> Optional[str]:
	soup = BeautifulSoup(html, "html.parser")
	meta = soup.select_one('meta[name="csrf-token"]')
	if meta and meta.get("content"):
		return meta["content"]
	hidden = soup.select_one('input[name="_token"]')
	if hidden and hidden.get("value"):
		return hidden["value"]
	return None


def login(session: requests.Session) -> None:
	email = os.getenv("IAC_EMAIL")
	password = os.getenv("IAC_PASSWORD")
	if not email or not password:
		raise RuntimeError("Missing IAC_EMAIL / IAC_PASSWORD in environment or .env")

	r = session.get(LOGIN_URL, timeout=25)
	r.raise_for_status()
	csrf = _extract_csrf_from_html(r.text)
	if not csrf:
		raise RuntimeError("Could not find CSRF token on login page.")

	payload = {"email": email, "password": password, "_token": csrf}
	r2 = session.post(LOGIN_URL, data=payload, timeout=35, allow_redirects=True)
	r2.raise_for_status()

	probe = session.get(INTERNAL_PROBE_URL, timeout=25, allow_redirects=True)
	if probe.status_code != 200 or "/login" in probe.url:
		raise RuntimeError("Login appears to have failed.")


def _parse_int(val: str) -> Optional[int]:
	try:
		return int(val)
	except Exception:
		return None


def load_primary_centers(csv_path: str, max_centers: int) -> List[Center]:
	centers: List[Center] = []
	with open(csv_path, "r", encoding="utf-8", newline="") as f:
		reader = csv.DictReader(f)
		for row in reader:
			status = (row.get("status_cd") or "").strip().lower()
			parent_id = (row.get("parent_id") or "").strip()
			symbol = (row.get("symbol") or "").strip()
			name = (row.get("name") or "").strip()
			center_id = _parse_int((row.get("center_id") or "").strip())

			if not center_id or not symbol or status != "active":
				continue

			# Primary centers have no parent_id (NULL or empty)
			if parent_id and parent_id.upper() != "NULL":
				continue

			centers.append(Center(center_id=center_id, symbol=symbol, name=name))

	centers.sort(key=lambda c: c.center_id)
	return centers[:max_centers]


def _safe_filename(name: str) -> str:
	name = re.sub(r"[^a-zA-Z0-9._-]+", "_", name).strip("_")
	return name or "activity_log"


def load_assigned_ids_from_sql(sql_path: str) -> set[str]:
	assigned_ids: set[str] = set()
	if not sql_path:
		return assigned_ids

	path = Path(sql_path)
	if not path.exists():
		print(f"SQL file not found: {path}", file=sys.stderr)
		return assigned_ids

	text = path.read_text(encoding="utf-8", errors="ignore")
	# assigned_id is the first value in each VALUES tuple
	for m in re.finditer(r"\(\s*['\"]([A-Za-z0-9_-]+)['\"]\s*,", text):
		assigned_ids.add(m.group(1).strip().upper())

	return assigned_ids


def _find_header_row(rows: Iterable[Sequence[object]], max_scan: int = 10) -> Optional[int]:
	for idx, row in enumerate(rows, start=1):
		if idx > max_scan:
			break
		headers = [str(c).strip().lower() if c is not None else "" for c in row]
		if any(h in ("assigned_id", "assigned id", "assessment id", "id") for h in headers):
			return idx
	return None


def extract_assigned_ids_from_excel(xlsx_path: Path) -> List[str]:
	wb = load_workbook(filename=xlsx_path, read_only=True, data_only=True)
	ws = wb.active

	rows = list(ws.iter_rows(values_only=True, max_row=10))
	header_row_idx = _find_header_row(rows)
	if header_row_idx is None:
		wb.close()
		return []

	header_row = rows[header_row_idx - 1]
	header_map = {
		(str(h).strip().lower() if h is not None else ""): i
		for i, h in enumerate(header_row)
	}

	col_idx = None
	for key in ("assigned_id", "assigned id", "assessment id", "id"):
		if key in header_map:
			col_idx = header_map[key]
			break

	if col_idx is None:
		wb.close()
		return []

	ids: List[str] = []
	for row in ws.iter_rows(values_only=True, min_row=header_row_idx + 1):
		if not row or col_idx >= len(row):
			continue
		val = row[col_idx]
		if val is None:
			continue
		sval = str(val).strip().upper()
		if sval:
			ids.append(sval)

	wb.close()
	return ids


def write_new_rows_csv(xlsx_path: Path, assigned_ids: set[str], out_csv: Path) -> int:
	wb = load_workbook(filename=xlsx_path, read_only=True, data_only=True)
	ws = wb.active

	rows = list(ws.iter_rows(values_only=True, max_row=10))
	header_row_idx = _find_header_row(rows)
	if header_row_idx is None:
		wb.close()
		return 0

	header_row = rows[header_row_idx - 1]
	headers = [str(h).strip() if h is not None else "" for h in header_row]
	header_map = {
		(str(h).strip().lower() if h is not None else ""): i
		for i, h in enumerate(header_row)
	}

	col_idx = None
	for key in ("assigned_id", "assigned id", "assessment id", "id"):
		if key in header_map:
			col_idx = header_map[key]
			break

	if col_idx is None:
		wb.close()
		return 0

	out_csv.parent.mkdir(parents=True, exist_ok=True)
	count = 0
	with open(out_csv, "w", encoding="utf-8", newline="") as f:
		writer = csv.writer(f)
		writer.writerow(headers)
		for row in ws.iter_rows(values_only=True, min_row=header_row_idx + 1):
			if not row or col_idx >= len(row):
				continue
			val = row[col_idx]
			if val is None:
				continue
			sval = str(val).strip().upper()
			if not sval or sval in assigned_ids:
				continue
			writer.writerow(list(row))
			count += 1

	wb.close()
	return count


def download_center_excel(session: requests.Session, center: Center, out_dir: Path) -> Path:
	url = DOWNLOAD_URL.format(symbol=center.symbol)
	r = session.get(url, timeout=60, stream=True)
	if r.status_code in (401, 403):
		raise DownloadError(f"Unauthorized for {center.symbol}")
	r.raise_for_status()

	filename = None
	cd = r.headers.get("Content-Disposition") or r.headers.get("content-disposition")
	if cd:
		m = re.search(r'filename="?([^";]+)"?', cd)
		if m:
			filename = m.group(1)

	if not filename:
		filename = f"{center.symbol}_activity_log.xlsx"

	filename = _safe_filename(filename)
	out_path = out_dir / filename

	with open(out_path, "wb") as f:
		for chunk in r.iter_content(chunk_size=1024 * 64):
			if chunk:
				f.write(chunk)

	return out_path


def main() -> None:
	csv_path = Path(CENTER_CSV)
	if not csv_path.exists():
		print(f"Center CSV not found: {csv_path}", file=sys.stderr)
		return

	out_dir = Path(OUT_DIR)
	out_dir.mkdir(parents=True, exist_ok=True)

	centers = load_primary_centers(str(csv_path), MAX_CENTERS)
	if not centers:
		print("No centers found to download.", file=sys.stderr)
		return

	assigned_ids = load_assigned_ids_from_sql(SQL_PATH)
	if assigned_ids:
		print(f"Loaded {len(assigned_ids)} assigned_id values from {SQL_PATH}.")
	else:
		print("No assigned_id values loaded (SQL file missing or empty).")

	session = build_session()
	try:
		login(session)
	except Exception as e:
		print(f"Login failed: {e}", file=sys.stderr)
		return

	print(f"Downloading {len(centers)} activity logs to {out_dir}...")

	for idx, center in enumerate(centers, start=1):
		try:
			out_path = download_center_excel(session, center, out_dir)
			print(f"[{idx}/{len(centers)}] {center.symbol} - saved {out_path.name}")

			if assigned_ids:
				excel_ids = extract_assigned_ids_from_excel(out_path)
				if not excel_ids:
					print(f"[{idx}/{len(centers)}] {center.symbol} - no assigned_id column found")
				else:
					dupes = [x for x in excel_ids if x in assigned_ids]
					if dupes:
						print(
							f"[{idx}/{len(centers)}] {center.symbol} - {len(dupes)}/{len(excel_ids)} already exist in SQL"
						)
						out_csv = out_dir / f"{center.symbol}_activity_log_new.csv"
						new_count = write_new_rows_csv(out_path, assigned_ids, out_csv)
						print(
							f"[{idx}/{len(centers)}] {center.symbol} - wrote {new_count} new rows to {out_csv.name}"
						)
					else:
						print(f"[{idx}/{len(centers)}] {center.symbol} - no duplicates found")
		except Exception as e:
			print(f"[{idx}/{len(centers)}] {center.symbol} - error: {e}", file=sys.stderr)
		time.sleep(REQUEST_DELAY_SEC)


if __name__ == "__main__":
	main()
