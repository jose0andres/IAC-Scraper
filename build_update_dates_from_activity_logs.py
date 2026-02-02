#!/usr/bin/env python3
"""
build_update_dates_from_activity_logs.py

Generate SQL updates to fill NULL assessment_upload_date and
implementation_upload_date using activity-log XLSX exports.

It matches rows by assigned_id (Excel column: ID).

Environment variables (optional):
  INPUT_DIR=downloads/activity_logs
    INPUT_GLOB=IAC_Assessments_*.xlsx
  OUT_SQL=update_upload_impl_dates.sql
  TABLE_NAME=assessment
    SQL_PATH=004_assessment_data.sql

Usage:
  python build_update_dates_from_activity_logs.py
"""

import csv
import os
import re
import sys
from pathlib import Path
from typing import Dict, Optional

from openpyxl import load_workbook

INPUT_DIR = os.getenv("INPUT_DIR", "downloads/activity_logs").strip()
INPUT_GLOB = os.getenv("INPUT_GLOB", "IAC_Assessments_*.xlsx").strip()
OUT_SQL = os.getenv("OUT_SQL", "update_upload_impl_dates.sql").strip()
TABLE_NAME = os.getenv("TABLE_NAME", "assessment").strip()
SQL_PATH = os.getenv("SQL_PATH", "004_assessment_data.sql").strip()

HEADER_ALIASES = {
    "id": "assigned_id",
    "upload completed": "assessment_upload_date",
    "implementation completed": "implementation_upload_date",
}


def _normalize_header_key(h: str) -> str:
    return " ".join(h.strip().lower().replace("_", " ").split())


def _build_header_map(headers: list[str]) -> Dict[str, int]:
    header_map: Dict[str, int] = {}
    for idx, h in enumerate(headers):
        key = _normalize_header_key(h)
        canonical = HEADER_ALIASES.get(key, key)
        header_map[canonical] = idx
    return header_map


def _read_xlsx(path: Path):
    wb = load_workbook(filename=path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return None, []
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    data_rows = [list(r) for r in rows[1:]]
    return headers, data_rows


def _normalize_datetime(val: Optional[str]) -> Optional[str]:
    if val is None:
        return None
    s = str(val).strip()
    if not s:
        return None

    # Already ISO-ish
    if re.match(r"^\d{4}-\d{2}-\d{2}(\s+\d{2}:\d{2}:\d{2})?$", s):
        return s if " " in s else f"{s} 00:00:00"

    # MM/DD/YYYY or MM/DD/YYYY HH:MM(:SS)
    m = re.match(
        r"^(?P<m>\d{1,2})/(?P<d>\d{1,2})/(?P<y>\d{4})(?:\s+(?P<h>\d{1,2}):(?P<mi>\d{2})(?::(?P<s>\d{2}))?)?$",
        s,
    )
    if m:
        y = int(m.group("y"))
        mo = int(m.group("m"))
        d = int(m.group("d"))
        h = int(m.group("h") or 0)
        mi = int(m.group("mi") or 0)
        sec = int(m.group("s") or 0)
        return f"{y:04d}-{mo:02d}-{d:02d} {h:02d}:{mi:02d}:{sec:02d}"

    return s


def _sql_value(val: Optional[str]) -> str:
    if val is None:
        return "NULL"
    s = str(val).strip()
    if s == "":
        return "NULL"
    s = s.replace("\\", "\\\\").replace("'", "''")
    return f"'{s}'"


def _parse_column_list(sql_text: str) -> Optional[list[str]]:
    m = re.search(r"INSERT\s+INTO\s+assessment\s*\((.*?)\)\s*VALUES", sql_text, re.IGNORECASE | re.DOTALL)
    if not m:
        return None
    cols_raw = m.group(1)
    cols = [c.strip().strip("`") for c in cols_raw.split(",")]
    return cols


def _parse_values_tuples(values_text: str) -> list[list[str]]:
    tuples: list[list[str]] = []
    current: list[str] = []
    buf: list[str] = []
    in_string = False
    escape = False
    depth = 0

    for ch in values_text:
        if in_string:
            buf.append(ch)
            if escape:
                escape = False
            elif ch == "\\":
                escape = True
            elif ch == "'":
                in_string = False
            continue

        if ch == "'":
            in_string = True
            buf.append(ch)
            continue

        if ch == "(":
            depth += 1
            if depth == 1:
                current = []
                buf = []
                continue
        elif ch == ")":
            if depth == 1:
                value = "".join(buf).strip()
                current.append(value)
                tuples.append(current)
                buf = []
                current = []
                depth -= 1
                continue
            depth -= 1
        elif ch == "," and depth == 1:
            value = "".join(buf).strip()
            current.append(value)
            buf = []
            continue

        if depth >= 1:
            buf.append(ch)

    return tuples


def _clean_sql_value(raw: str) -> Optional[str]:
    if raw is None:
        return None
    s = raw.strip()
    if s.upper() == "NULL" or s == "":
        return None
    if s.startswith("'") and s.endswith("'"):
        s = s[1:-1]
        s = s.replace("''", "'")
    return s


def _load_existing_dates(sql_path: str) -> Dict[str, tuple[Optional[str], Optional[str]]]:
    path = Path(sql_path)
    if not path.exists():
        print(f"SQL file not found: {path}", file=sys.stderr)
        return {}

    sql_text = path.read_text(encoding="utf-8", errors="ignore")
    cols = _parse_column_list(sql_text)
    if not cols:
        print("Could not parse column list from SQL file.", file=sys.stderr)
        return {}

    try:
        assigned_idx = cols.index("assigned_id")
        upload_idx = cols.index("assessment_upload_date")
        impl_idx = cols.index("implementation_upload_date")
    except ValueError:
        print("Required columns not found in SQL file.", file=sys.stderr)
        return {}

    values_start = re.search(r"VALUES", sql_text, re.IGNORECASE)
    if not values_start:
        return {}
    values_text = sql_text[values_start.end():]

    rows = _parse_values_tuples(values_text)
    existing: Dict[str, tuple[Optional[str], Optional[str]]] = {}
    for row in rows:
        if len(row) <= max(assigned_idx, upload_idx, impl_idx):
            continue
        assigned_id = _clean_sql_value(row[assigned_idx])
        if not assigned_id:
            continue
        upload_val = _clean_sql_value(row[upload_idx])
        impl_val = _clean_sql_value(row[impl_idx])
        existing[assigned_id] = (upload_val, impl_val)

    return existing


def main() -> None:
    existing_dates = _load_existing_dates(SQL_PATH)
    if not existing_dates:
        print("No existing dates loaded from SQL file.", file=sys.stderr)

    input_dir = Path(INPUT_DIR)
    if not input_dir.exists():
        print(f"Input directory not found: {input_dir}", file=sys.stderr)
        return

    files = sorted(input_dir.glob(INPUT_GLOB))
    if not files:
        print(f"No XLSX files found in {input_dir} matching {INPUT_GLOB}")
        return

    out_path = Path(OUT_SQL)
    if not out_path.is_absolute():
        out_path = input_dir / out_path

    updates = 0

    with open(out_path, "w", encoding="utf-8", newline="") as out_f:
        out_f.write("use itac_ap;\n")

        for file_path in files:
            headers, rows = _read_xlsx(file_path)
            if not headers:
                continue

            current_map = _build_header_map(headers)
            if "assigned_id" not in current_map:
                print(f"Missing ID column in {file_path.name}", file=sys.stderr)
                return

            for row in rows:
                assigned_id = None
                if current_map["assigned_id"] < len(row):
                    assigned_id = str(row[current_map["assigned_id"]]).strip()

                    if not assigned_id:
                        continue

                    upload_val = None
                    if "assessment_upload_date" in current_map and current_map["assessment_upload_date"] < len(row):
                        upload_val = _normalize_datetime(row[current_map["assessment_upload_date"]])

                    impl_val = None
                    if "implementation_upload_date" in current_map and current_map["implementation_upload_date"] < len(row):
                        impl_val = _normalize_datetime(row[current_map["implementation_upload_date"]])

                    if not upload_val and not impl_val:
                        continue

                    existing_upload, existing_impl = existing_dates.get(assigned_id, (None, None))

                    if existing_upload is not None and existing_impl is not None:
                        continue

                    if existing_upload is not None:
                        upload_val = None
                    if existing_impl is not None:
                        impl_val = None

                    if not upload_val and not impl_val:
                        continue

                    out_f.write(
                        f"UPDATE {TABLE_NAME}\n"
                        f"SET assessment_upload_date = IFNULL(assessment_upload_date, {_sql_value(upload_val)}),\n"
                        f"    implementation_upload_date = IFNULL(implementation_upload_date, {_sql_value(impl_val)})\n"
                        f"WHERE assigned_id = {_sql_value(assigned_id)};\n"
                    )
                    updates += 1


    print(f"Wrote {updates} updates to {out_path}")


if __name__ == "__main__":
    main()
