import argparse
import datetime as dt
from pathlib import Path
from typing import List

from openpyxl import load_workbook
from openpyxl.workbook import Workbook


def _list_input_files(input_dir: Path, pattern: str) -> List[Path]:
    return sorted(input_dir.glob(pattern))


def _read_headers(xlsx_path: Path) -> List[str]:
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        ws = wb.active
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        return [str(h).strip() if h is not None else "" for h in header_row]
    finally:
        wb.close()


def _build_master_headers(files: List[Path]) -> List[str]:
    master: List[str] = []
    seen = set()
    for f in files:
        headers = _read_headers(f)
        for h in headers:
            if h and h not in seen:
                master.append(h)
                seen.add(h)
    return master


def _row_is_empty(row: List[object]) -> bool:
    for v in row:
        if v not in (None, ""):
            return False
    return True


def combine_workbooks(
    input_dir: Path,
    pattern: str,
    output_path: Path,
    add_source: bool,
) -> int:
    files = _list_input_files(input_dir, pattern)
    if not files:
        raise FileNotFoundError(f"No files found matching {pattern} in {input_dir}")

    master_headers = _build_master_headers(files)
    if add_source:
        master_headers = master_headers + ["source_file"]

    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.title = "combined"
    out_ws.append(master_headers)

    total_rows = 0
    for f in files:
        wb = load_workbook(f, read_only=True, data_only=True)
        try:
            ws = wb.active
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            headers = [str(h).strip() if h is not None else "" for h in header_row]
            index_map = {h: idx for idx, h in enumerate(headers) if h}

            for row in ws.iter_rows(min_row=2, values_only=True):
                if row is None:
                    continue
                if _row_is_empty(list(row)):
                    continue

                out_row = []
                for h in master_headers:
                    if add_source and h == "source_file":
                        out_row.append(f.name)
                        continue
                    idx = index_map.get(h)
                    out_row.append(row[idx] if idx is not None and idx < len(row) else None)

                out_ws.append(out_row)
                total_rows += 1
        finally:
            wb.close()

    output_path.parent.mkdir(parents=True, exist_ok=True)
    out_wb.save(output_path)
    return total_rows


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Combine IAC_Assessments Excel files into one workbook."
    )
    parser.add_argument(
        "--input-dir",
        default=str(Path("downloads") / "activity_logs"),
        help="Directory to search for input files.",
    )
    parser.add_argument(
        "--pattern",
        default="IAC_Assessments*.xlsx",
        help="Glob pattern for input files.",
    )
    parser.add_argument(
        "--output",
        default=None,
        help="Output .xlsx path. Defaults to downloads/activity_logs/IAC_Assessments_combined_YYYYMMDD.xlsx",
    )
    parser.add_argument(
        "--add-source",
        action="store_true",
        help="Add source_file column with the original filename.",
    )

    args = parser.parse_args()
    input_dir = Path(args.input_dir)

    if args.output:
        output_path = Path(args.output)
    else:
        today = dt.datetime.now().strftime("%Y%m%d")
        output_path = (
            Path("downloads")
            / "activity_logs"
            / f"IAC_Assessments_combined_{today}.xlsx"
        )

    total_rows = combine_workbooks(
        input_dir=input_dir,
        pattern=args.pattern,
        output_path=output_path,
        add_source=args.add_source,
    )

    print(f"Combined {total_rows} rows into {output_path}")


if __name__ == "__main__":
    main()
